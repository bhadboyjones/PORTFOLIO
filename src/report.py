"""
report.py
=========
Excel report generation for BTM BESS scenario results.

Expects a list of scenario DataFrames (or pickle file paths) that have been
through the full pipeline:

    build_optimiser_input()   — data_builder
    calculate_baseline()      — baseline site cost without BESS
    run_optimiser()           — BESS dispatch
    calculate_settlement()    — BESS P&L + standing charges

Each DataFrame must contain:
    scenario_label      str column — BESS config label (e.g. "1MW_2h")
    export_limit_mw     float column — export cap for this scenario
    site_name           str column — site archetype name (optional)
    All baseline_*, net_settlement_gbp, total_standing_gbp, dispatch columns

Call:
    from src.report import build_report
    build_report(all_results, "../outputs/bess_scenarios_pnl.xlsx")

all_results may be a list of DataFrames or a list of pickle file paths (str).
When paths are passed, each file is loaded, used, then freed — only one
scenario DataFrame is in memory at any time.
"""

import os
from datetime import datetime, timezone

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_BODY_FONT      = Font(name="Arial", size=10)

_FILL_DARK_BLUE = PatternFill("solid", start_color="1F4E79")
_FILL_MID_BLUE  = PatternFill("solid", start_color="2E75B6")

_CENTER = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_float(value):
    """Round floats for Excel; pass strings and ints through unchanged."""
    if isinstance(value, float):
        return round(value, 2)
    return value


def _styled_cell(ws, value, fill, font=None):
    """Return a single styled WriteOnlyCell."""
    cell = WriteOnlyCell(ws, value=value)
    cell.font      = font or _HEADER_FONT
    cell.fill      = fill
    cell.alignment = _CENTER
    return cell


def _header_row(ws, col_names, fill=None):
    """Return a list of styled WriteOnlyCell objects for a header row."""
    fill = fill or _FILL_DARK_BLUE
    return [_styled_cell(ws, name, fill) for name in col_names]


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

# Section definitions: (n_cols, label, fill)
_SUMMARY_SECTIONS = [
    (3,  "Site",              _FILL_DARK_BLUE),
    (4,  "Baseline (no BESS)", _FILL_MID_BLUE),
    (4,  "Standing Charges",   _FILL_MID_BLUE),
    (6,  "BESS P&L",           _FILL_MID_BLUE),
    (3,  "Comparison",         _FILL_MID_BLUE),
]


def _build_summary_row(scenario):
    """Build one summary row dict for a single scenario DataFrame."""
    baseline_net   = scenario["baseline_net_gbp"].sum()
    total_standing = scenario["total_standing_gbp"].sum()
    net_settlement = scenario["net_settlement_gbp"].sum()

    site_cost_wo_bess = baseline_net + total_standing
    site_cost_w_bess  = site_cost_wo_bess - net_settlement
    bess_net_benefit  = net_settlement

    return {
        "site_name":                scenario["site_name"].iloc[0] if "site_name" in scenario.columns else "",
        "scenario_label":           scenario["scenario_label"].iloc[0],
        "export_limit_mw":          scenario["export_limit_mw"].iloc[0],
        "baseline_import_cost_gbp": scenario["baseline_import_cost_gbp"].sum(),
        "baseline_export_rev_gbp":  scenario["baseline_export_rev_gbp"].sum(),
        "baseline_chp_cost_gbp":    scenario["baseline_chp_cost_gbp"].sum(),
        "baseline_net_gbp":         baseline_net,
        "dduos_fixed_gbp":          scenario["dduos_fixed_gbp"].sum(),
        "dduos_capacity_gbp":       scenario["dduos_capacity_gbp"].sum(),
        "gduos_fixed_gbp":          scenario["gduos_fixed_gbp"].sum(),
        "total_standing_gbp":       total_standing,
        "dis1_saving_gbp":          scenario["dis1_saving_gbp"].sum(),
        "dis2_revenue_gbp":         scenario["dis2_revenue_gbp"].sum(),
        "charge2_cost_gbp":         scenario["charge2_cost_gbp"].sum(),
        "charge1_opp_cost_gbp":     scenario["charge1_opp_cost_gbp"].sum(),
        "deg_cost_gbp":             scenario["deg_cost_gbp"].sum(),
        "net_settlement_gbp":       net_settlement,
        "site_cost_wo_bess_gbp":    site_cost_wo_bess,
        "site_cost_w_bess_gbp":     site_cost_w_bess,
        "bess_net_benefit_gbp":     bess_net_benefit,
    }


def _write_summary_sheet(wb, summary_df):
    """
    Write the Summary sheet in write_only mode.

    merge_cells is not available in write_only mode, so section labels occupy
    the first cell of each group with the same fill spanning the remaining
    cells as empty styled cells — visually equivalent.
    """
    ws = wb.create_sheet(title="Summary")
    ws.sheet_format.defaultColWidth = 26

    # Row 1: section labels
    section_row = []
    for n_cols, label, fill in _SUMMARY_SECTIONS:
        section_row.append(_styled_cell(ws, label, fill))
        for _ in range(n_cols - 1):
            section_row.append(_styled_cell(ws, None, fill))
    ws.append(section_row)

    # Row 2: column headers
    ws.append(_header_row(ws, list(summary_df.columns)))

    # Data rows
    for row in summary_df.itertuples(index=False):
        ws.append([_safe_float(v) for v in row])

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Detail sheets
# ---------------------------------------------------------------------------

_DETAIL_COLS = [
    "startTime",
    "net_demand_mw",
    "net_demand_mwh",
    "baseline_import_cost_gbp",
    "baseline_export_rev_gbp",
    "baseline_chp_cost_gbp",
    "baseline_net_gbp",
    "dduos_fixed_gbp",
    "dduos_capacity_gbp",
    "gduos_fixed_gbp",
    "total_standing_gbp",
    "charge1_mw",
    "charge2_mw",
    "dis1_mw",
    "dis2_mw",
    "soc_mwh",
    "dis1_saving_gbp",
    "dis2_revenue_gbp",
    "charge2_cost_gbp",
    "charge1_opp_cost_gbp",
    "deg_cost_gbp",
    "net_settlement_gbp",
    "import_rate_gbp",
    "export_rate_gbp",
]


def _write_detail_sheet(wb, scenario):
    """Stream one SP-level detail sheet per scenario using ws.append()."""
    label        = scenario["scenario_label"].iloc[0]
    export_limit = scenario["export_limit_mw"].iloc[0]
    site         = scenario["site_name"].iloc[0] if "site_name" in scenario.columns else ""
    sheet_name   = f"{site}_{label}_exp{export_limit}".replace(".", "p")[:31]

    scenario = scenario.copy()
    scenario["startTime"] = pd.to_datetime(scenario["startTime"]).dt.tz_localize(None)

    ws   = wb.create_sheet(title=sheet_name)
    ws.sheet_format.defaultColWidth = 26
    cols = [c for c in _DETAIL_COLS if c in scenario.columns]

    ws.append(_header_row(ws, cols))

    for row in scenario[cols].itertuples(index=False):
        ws.append([_safe_float(v) for v in row])

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def _write_metadata_sheet(wb, job_id):
    """Write a Metadata sheet as the first tab — makes the file self-identifying."""
    ws = wb.create_sheet(title="Metadata")
    ws.sheet_format.defaultColWidth = 36

    ws.append(_header_row(ws, ["Field", "Value"]))
    ws.append(["Job ID",    job_id])
    ws.append(["Generated", datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")])


def build_report(all_results, output_path, job_id=None):
    """
    Build and save the full BESS scenario Excel report.

    Uses openpyxl write_only mode — each row is streamed to disk as it is
    appended, so the workbook never holds more than one row's worth of data
    in memory regardless of how many scenarios or SP rows are written.

    Args:
        all_results  : list of scenario DataFrames or pickle file paths (str).
                       When paths are given, each is loaded, processed, and
                       freed before the next is loaded.
        output_path  : destination .xlsx path (used as the output directory;
                       filename is replaced with bess_results_{job_id}.xlsx
                       when job_id is provided)
        job_id       : optional run identifier (YYYYMMDD_HHMMSS_xxxx).
                       When provided: appended to the output filename and
                       written to a Metadata sheet as the first tab.

    Returns:
        str — the path the file was actually written to.

    Output sheets (in order):
        Metadata     : job_id + generated timestamp (only when job_id provided)
        Summary      : one row per scenario, full P&L stack, sorted by net benefit
        <scenario>   : one SP-level detail sheet per scenario
    """
    if not all_results:
        raise ValueError("all_results is empty — nothing to report.")

    # Derive final output path
    if job_id is not None:
        output_path = os.path.join(
            os.path.dirname(output_path),
            f"bess_results_{job_id}.xlsx",
        )

    def _load(item):
        if isinstance(item, str):
            return pd.read_pickle(item)
        return item

    # Pass 1: collect summary rows — one DF loaded at a time, freed each iteration
    summary_rows = []
    for item in all_results:
        scenario = _load(item)
        summary_rows.append(_build_summary_row(scenario))

    summary_df = pd.DataFrame(summary_rows).sort_values(
        "bess_net_benefit_gbp", ascending=False
    )

    # Write workbook in streaming mode — no in-memory cell tree
    wb = Workbook(write_only=True)

    # Metadata sheet first (tab 1) when job_id is present
    if job_id is not None:
        _write_metadata_sheet(wb, job_id)

    # Summary sheet next
    _write_summary_sheet(wb, summary_df)

    # Pass 2: stream detail sheets — one DF in memory at a time
    for item in all_results:
        scenario = _load(item)
        _write_detail_sheet(wb, scenario)

    wb.save(output_path)
    print(f"Report saved — {len(all_results)} scenarios → {output_path}")

    return output_path
